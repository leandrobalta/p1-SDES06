from utils.json import read_json_data
from models.discipline import Discipline
from models.graph import Graph
from utils.format import format_schedule
from utils.color_assignment import color_function
import networkx as nx
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import Alignment, Font
from typing import List, Dict


def get_most_teacher_classes(graph: Graph):
    teacher_count = {}

    for node in graph.graph.nodes:
        discipline = graph.graph.nodes[node]["discipline"]
        teacher_count[discipline.teacher] = teacher_count.get(discipline.teacher, 0) + 1

    max_count = max(teacher_count.values())
    max_teachers = [
        teacher for teacher, count in teacher_count.items() if count == max_count
    ]

    return max_teachers


def get_most_heavy_semester(graph: Graph):
    semester_count = {}

    for node in graph.graph.nodes:
        discipline = graph.graph.nodes[node]["discipline"]
        semester_count[discipline.semester] = (
            semester_count.get(discipline.semester, 0) + 1
        )

    max_count = max(semester_count.values())
    max_semesters = [
        semester for semester, count in semester_count.items() if count == max_count
    ]

    return max_semesters


def define_weight(graph: Graph):
    def sin_discipline(discipline: Discipline):
        return discipline.course.lower() == "sin"

    def cco_discipline(discipline: Discipline):
        return discipline.course.lower() == "cco"

    def most_teacher_classes(discipline: Discipline):
        return discipline.teacher in get_most_teacher_classes(graph)

    def most_heavy_semester(discipline: Discipline):
        return discipline.semester in get_most_heavy_semester(graph)

    for node in graph.graph.nodes:
        discipline = graph.graph.nodes[node]["discipline"]
        weight = 0
        if sin_discipline(discipline):
            weight += 20
        if cco_discipline(discipline):
            weight += 10
        if most_teacher_classes(discipline):
            weight += 3
        if most_heavy_semester(discipline):
            weight += 2
        discipline.set_weight(weight)

    graph.order_by_weight()
    return graph


def evaluate_schedule_score(
    graph: Graph, slots: list, slot: int, discipline: Discipline, day: str
):
    score = 0

    # **Penalidades**

    if day.lower() == "saturday" and discipline.course.lower() in ["sin", "cco"]:
        return float("-inf")

    # Penaliza por conflito de professor ou semestre/curso
    for existing_discipline in slots[slot]:
        if graph.graph.has_edge(discipline.index, existing_discipline["index"]):
            return float("-inf")

    # Penaliza se já houverem 8 ou mais aulas do curso no semestre naquele dia
    count = 0
    for day_slots in slots:
        for existing_discipline in day_slots:
            if (
                discipline.semester == existing_discipline["semester"]
                and discipline.course == existing_discipline["course"]
            ):
                count += 1

    if count >= 8:
        score -= 3

    # Penaliza se a disciplina já tiver o limite de aulas por dia
    count = 0
    for day_slots in slots:
        for existing_discipline in day_slots:
            if discipline.index == existing_discipline["index"]:
                count += 1

            if count >= 2 and discipline.ch == 4:
                score -= 8
            elif count >= 3 and discipline.ch == 5:
                score -= 8

    # Penaliza se o professor tiver mais de 6 aulas no dia
    count = 0
    for day_slots in slots:
        for existing_discipline in day_slots:
            if existing_discipline["teacher"] == discipline.teacher:
                count += 1

            if count >= 6:
                score -= 2

    # Penaliza se o slot já tiver o limite máximo de disciplinas (8)
    if len(slots[slot]) >= 8:
        score -= 3

    # **Benefícios**

    # Adiciona pontos para slot vazio
    if len(slots[slot]) == 0:
        score += 3

    # Adiciona pontos para continuidade de horário (se disciplina requer mais de uma aula seguida)
    if discipline.ch > 2:
        if slot > 0 and any(
            discipline_in_slot["index"] == discipline.index
            for discipline_in_slot in slots[slot - 1]
        ):
            score += 6
        if slot < len(slots) - 1 and any(
            discipline_in_slot["index"] == discipline.index
            for discipline_in_slot in slots[slot + 1]
        ):
            score += 6

    # Adiciona pontos para distribuição balanceada de aulas por professor no dia (menos de 5 aulas)
    if count < 5:
        score += 2

    # Adiciona pontos para slots que ajudam a distribuir o curso e semestre de forma balanceada na semana
    if count < 5:
        score += 1

    return score


def is_night_period(discipline: Discipline, slot: int):
    return discipline.course.lower() == "sin" and slot >= 10


def generate_schedule(graph: Graph):
    for node in graph.order_by_weight():
        discipline = graph.graph.nodes[node]["discipline"]
        slots_needed = discipline.ch

        # Loop até alocar todos os slots necessários para a carga horária
        while slots_needed > 0:
            best_slot_sequence = None
            best_score = float("-inf")

            for day_schedule in graph.schedules.get_schedule():
                for day, slots in day_schedule.items():
                    for slot in range(len(slots)):

                        # Verifica restrições de horários
                        if discipline.course.lower() == "sin" and not is_night_period(
                            discipline, slot
                        ):
                            continue

                        # Avalia o slot atual e calcula a pontuação
                        score = evaluate_schedule_score(
                            graph, slots, slot, discipline, day
                        )

                        # Se a pontuação é a melhor até agora, salva o slot
                        if score > best_score:
                            best_score = score
                            best_slot_sequence = (day, slots, slot)

            # Aloca o slot se encontrou uma posição adequada
            if best_slot_sequence and best_score > 0:
                day, slots, slot = best_slot_sequence
                slots[slot].append(
                    {
                        "name": discipline.name,
                        "code": discipline.code,
                        "teacher": discipline.teacher,
                        "course": discipline.course,
                        "semester": discipline.semester,
                        "ch": discipline.ch,
                        "index": discipline.index,
                    }
                )
                discipline.add_schedule({day: slot})
                slots_needed -= 1
            else:
                # Se não há slots disponíveis para a disciplina, interrompe o loop
                break
    return graph


def generate_edge(graph: Graph):
    for node1 in graph.graph.nodes:
        discipline1 = graph.graph.nodes[node1]["discipline"]
        for node2 in graph.graph.nodes:
            if node1 != node2:
                discipline2 = graph.graph.nodes[node2]["discipline"]
                if (discipline1.teacher == discipline2.teacher) or (
                    discipline1.semester == discipline2.semester
                    and discipline1.course == discipline2.course
                ):
                    graph.add_edge(node1, node2)
    return graph


def generate_graph(data: list):
    graph = Graph()
    count = 0
    for index, discipline_data in enumerate(data):
        count += discipline_data["ch"]
        discipline = Discipline(
            index,
            discipline_data["name"],
            discipline_data["teacher"],
            discipline_data["code"],
            discipline_data["semester"],
            discipline_data["ch"],
            discipline_data["course"],
        )
        graph.add_node(discipline)

    graph = generate_edge(graph)
    print(f"Total classes: {count}")
    return graph


def plot_graph(graph: Graph):

    color_map = {}

    for node in graph.graph.nodes:
        discipline = graph.graph.nodes[node]["discipline"]

        if not discipline.schedule:
            color_map[node] = "gray"
        else:
            color_map[node] = color_function(str(discipline.schedule))

    colors = [color_map[node] for node in graph.graph.nodes]

    pos = nx.spring_layout(graph.graph)

    nx.draw(graph.graph, pos, node_color=colors, with_labels=True)

    plt.show()


def create_schedule_excel(graph, filename: str = "schedule.xlsx"):
    # Criação do workbook e planilhas
    workbook = openpyxl.Workbook()
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

    # Função para determinar o horário com base no índice do slot
    def get_slot_time(slot_index: int) -> str:
        if 0 <= slot_index <= 4:
            times = ["7:00", "7:55", "8:50", "10:00", "10:55"]
            return times[slot_index]
        elif 5 <= slot_index <= 9:
            times = ["13:00", "13:55", "14:50", "16:00", "16:55"]
            return times[slot_index - 5]
        else:
            times = ["19:00", "19:50", "21:00", "21:50", "22:40"]
            return times[(slot_index - 10) % len(times)]

    # Iterar pelos horários e disciplinas no cronograma
    for day_schedule in graph.schedules.get_schedule():
        for day, slots in day_schedule.items():
            # Criar uma nova aba para cada dia da semana
            sheet = workbook.create_sheet(title=day.capitalize())
            sheet.append(["Hour Slot", "Disciplines"])  # Cabeçalho
            sheet.column_dimensions["A"].width = 12
            sheet.column_dimensions["B"].width = 200

            for slot_index, disciplines in enumerate(slots):
                if disciplines:
                    discipline_info = "; ".join(
                        [discipline["name"] for discipline in disciplines]
                    )
                else:
                    discipline_info = "Free"

                # Adicionar dados à planilha
                sheet.append([get_slot_time(slot_index), discipline_info])

    # Remove a planilha padrão criada automaticamente
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]

    # Estilo do cabeçalho
    for sheet_name in days:
        if sheet_name.capitalize() in workbook.sheetnames:
            sheet = workbook[sheet_name.capitalize()]
            for cell in sheet[1]:  # Primeira linha
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Salvar o arquivo Excel
    workbook.save(filename)
    print(f"Schedule exported successfully to {filename}.")


def main():
    data = read_json_data("cenarios/cenario1.json")
    graph = generate_graph(data)
    graph = define_weight(graph)
    graph = generate_schedule(graph)
    # format_schedule(graph)
    # plot_graph(graph)
    # graph.print_disciplines()
    # graph.schedules.print_schedule()
    create_schedule_excel(graph)


if __name__ == "__main__":
    main()
